import openpyxl
from openpyxl.reader.excel import load_workbook
from russian_names import RussianNames

def generateFIO(amount:int) -> list:
    fio_list = []
    fio = ''
    for i in range(amount):
        fio = str(i + 1) + ' ' + RussianNames().get_person()
        fio_list.append(tuple(fio.split()))
    print(fio)
    print(fio_list)
    return fio_list

def clearWorksheet(sheet):
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.value = None

def writeToXlsx(write_list: list, ans: str):
    if ans == 'N' or ans == 'n':
        wb = load_workbook('people.xlsx')
        sheet = wb.active
        clearWorksheet(sheet)
        sheet.title = "People"
        last_row = 1
    elif ans == 'Y' or ans == 'y':
        try:
            wb = load_workbook('people.xlsx')
            sheet = wb.active
            last_row = sheet.max_row
        except FileNotFoundError:
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = "People"
            last_row = 1
    else:
        print('Error: Write N or Y only')
        return
    sheet['A1'] = "Номер"
    sheet['B1'] = "Фамилия"
    sheet['C1'] = "Имя"
    sheet['D1'] = "Отчество"
    for i, fio in enumerate(write_list):
            sheet.cell(row=last_row + i + 1, column=1).value = last_row + i
            sheet.cell(row=last_row + i + 1, column=2).value = fio[3]
            sheet.cell(row=last_row + i + 1, column=3).value = fio[1]
            sheet.cell(row=last_row + i + 1, column=4).value = fio[2]
    wb.save('people.xlsx')

def main():
    input_number = int(input('Enter a number: '))
    answer = input('Do you want to add FIOs to existing file? (N - no, Y - yes): ')
    fio_list = generateFIO(input_number)
    writeToXlsx(fio_list, answer)

if __name__ == "__main__":
    main()